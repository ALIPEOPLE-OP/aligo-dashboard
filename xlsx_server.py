#!/usr/bin/env python3
"""
xlsx_server.py — 로컬 Excel 파싱 서버
openpyxl로 xlsx 파일을 읽어 JSON으로 반환 (CDN 불필요)
"""
import json, sys, os
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs, urlparse
import io, cgi, traceback

PORT = 8765

try:
    import openpyxl
except ImportError:
    print("openpyxl 없음, 설치 중...")
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

def xlsx_to_json(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([
                str(cell) if cell is not None else ''
                for cell in row
            ])
        result[name] = rows
    return {"sheetNames": wb.sheetnames, "sheets": result}

class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args): pass  # 로그 억제

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        if self.path == '/health':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(b'{"status":"ok","version":"1.0"}')
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        if self.path != '/parse':
            self.send_response(404); self.end_headers(); return
        try:
            ctype = self.headers.get('Content-Type', '')
            if 'multipart/form-data' in ctype:
                form = cgi.FieldStorage(
                    fp=self.rfile, headers=self.headers,
                    environ={'REQUEST_METHOD': 'POST', 'CONTENT_TYPE': ctype}
                )
                item = form['file']
                data = item.file.read()
            else:
                length = int(self.headers.get('Content-Length', 0))
                data = self.rfile.read(length)

            result = xlsx_to_json(data)
            body = json.dumps(result, ensure_ascii=False, default=str).encode('utf-8')

            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)
        except Exception as e:
            err = json.dumps({"error": str(e)}).encode('utf-8')
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(err)

if __name__ == '__main__':
    server = HTTPServer(('127.0.0.1', PORT), Handler)
    print(f"xlsx_server 시작: http://127.0.0.1:{PORT}", flush=True)
    server.serve_forever()
